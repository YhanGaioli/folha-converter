import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import io
import json
import base64
from pathlib import Path
from http.server import BaseHTTPRequestHandler

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def limpar(t):
    return re.sub(r'\s*P[aá]gina:.*$', '', t).strip()

CAB="1F4E79"; PRO="E8F5E9"; DES="FFEBEE"; PRO2="C6EFCE"; DES2="FFC7CE"
INF="FFEB9C"; TOT="D9E1F2"; CC_COR="DEEAF1"

def ecab(c):
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", start_color=CAB)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    b = Side(style="thin", color="AAAAAA")
    c.border = Border(left=b, right=b, top=b, bottom=b)

def edat(c, cor=None):
    b = Side(style="thin", color="DDDDDD")
    c.border = Border(left=b, right=b, top=b, bottom=b)
    c.alignment = Alignment(vertical="center")
    if cor:
        c.fill = PatternFill("solid", start_color=cor)

# ─────────────────────────────────────────────
# DETECTOR DE TIPO
# ─────────────────────────────────────────────

def detectar_tipo(conteudo_bytes):
    with pdfplumber.open(io.BytesIO(conteudo_bytes)) as pdf:
        texto = pdf.pages[0].extract_text() or ""
    if "EXTRATO MENSAL" in texto:
        return "extrato"
    return "resumo"

# ─────────────────────────────────────────────
# PARSER: RESUMO POR CC
# ─────────────────────────────────────────────

def processar_resumo(conteudo_bytes, nome):
    registros = []
    cnpj = razao = comp = 'N/A'
    cc = 'Sem CC'
    tipo = None
    ignorar = False

    with pdfplumber.open(io.BytesIO(conteudo_bytes)) as pdf:
        for pagina in pdf.pages:
            for linha in (pagina.extract_text() or "").split('\n'):
                l = linha.strip()
                if not l: continue
                m = re.match(r'Empresa:\s*\d+\s*-\s*(.+)', l, re.I)
                if m: razao = limpar(m.group(1)); continue
                m = re.match(r'CNPJ:\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})', l, re.I)
                if m: cnpj = m.group(1); continue
                m = re.match(r'Compet.ncia:\s*(\d{2}/\d{4})', l, re.I)
                if m: comp = m.group(1); continue
                m = re.match(r'C\.?Custo:\s*(.+)', l, re.I)
                if m: cc = m.group(1).strip(); tipo = None; ignorar = False; continue
                if re.search(r'Resumo das bases', l, re.I): tipo = None; ignorar = True; continue
                if ignorar: continue
                if re.match(r'^(Folha Mensal|Total:|Liquido|Rubrica)\s*', l, re.I): continue
                if re.match(r'^PROVENTOS\s*$', l, re.I): tipo = 'Provento'; continue
                if re.match(r'^DESCONTOS\s*$', l, re.I): tipo = 'Desconto'; continue
                if re.match(r'^INFORMATIVA\s*$', l, re.I): tipo = 'Informativo'; continue
                m = re.match(r'^(\d{1,6})\s+(.+?)\s+[\d:]+\s+[\d:.,]+\s+([\d.]+,\d{2})\*?\s*$', l)
                if m and tipo:
                    v = float(m.group(3).replace('.', '').replace(',', '.'))
                    registros.append({
                        'CNPJ': cnpj, 'Razao Social': razao, 'Competencia': comp,
                        'Centro de Custo': cc, 'Tipo': tipo,
                        'Rubrica': m.group(1) + ' - ' + m.group(2).strip(),
                        'Valor': v, 'Arquivo': nome
                    })
    return registros

def gerar_excel_resumo(registros):
    df = pd.DataFrame(registros)
    df = df.groupby(['CNPJ','Razao Social','Competencia','Centro de Custo','Tipo','Rubrica'],
                    sort=False).agg(Valor=('Valor','sum')).reset_index()
    wb = Workbook(); wb.remove(wb.active)

    # Resumo
    ws = wb.create_sheet("RESUMO", 0)
    ws.merge_cells('A1:F1'); ws['A1'] = "RESUMO CONSOLIDADO - FOLHA DE PAGAMENTO"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws['A1'].alignment = Alignment(horizontal="center")
    for col,(h,w) in enumerate(zip(['CNPJ','Razao Social','Competencia','Total Proventos','Total Descontos','Total Informativos'],[22,42,14,18,18,18]),1):
        c = ws.cell(row=3, column=col, value=h); ecab(c)
        ws.column_dimensions[get_column_letter(col)].width = w
    linha = 4; tp=td=ti=0
    for (cnpj,comp),grp in df.groupby(['CNPJ','Competencia']):
        p=grp[grp.Tipo=='Provento'].Valor.sum()
        d=grp[grp.Tipo=='Desconto'].Valor.sum()
        i=grp[grp.Tipo=='Informativo'].Valor.sum()
        tp+=p; td+=d; ti+=i
        for col,val in enumerate([cnpj,grp['Razao Social'].iloc[0],comp,p,d,i],1):
            c=ws.cell(row=linha,column=col,value=val); edat(c)
            if col>=4: c.number_format='R$ #,##0.00'; c.alignment=Alignment(horizontal="right")
        linha+=1
    for col in range(1,7):
        c=ws.cell(row=linha,column=col); c.fill=PatternFill("solid",start_color=TOT); c.font=Font(bold=True)
    ws.merge_cells(f'A{linha}:C{linha}')
    ws.cell(row=linha,column=1,value="TOTAL GERAL").alignment=Alignment(horizontal="center")
    for col,val in zip([4,5,6],[tp,td,ti]):
        c=ws.cell(row=linha,column=col,value=val)
        c.number_format='R$ #,##0.00'; c.font=Font(bold=True); c.alignment=Alignment(horizontal="right")
    ws.freeze_panes='A4'

    def cor_tipo(t):
        return {'Provento':PRO2,'Desconto':DES2,'Informativo':INF}.get(t)

    for cnpj, dfc in df.groupby('CNPJ'):
        ws2 = wb.create_sheet(re.sub(r'[^\w]','',cnpj)[:31])
        razao=dfc['Razao Social'].iloc[0]; comp=dfc['Competencia'].iloc[0]
        ws2.merge_cells('A1:D1'); ws2['A1']=f'FOLHA DE PAGAMENTO - {razao}'
        ws2['A1'].font=Font(bold=True,size=13,color="1F4E79"); ws2['A1'].alignment=Alignment(horizontal="center")
        ws2.merge_cells('A2:D2'); ws2['A2']=f'CNPJ: {cnpj}  |  Competencia: {comp}'
        ws2['A2'].font=Font(italic=True,color="444444"); ws2['A2'].alignment=Alignment(horizontal="center")
        for col,(h,w) in enumerate(zip(['Centro de Custo','Tipo','Rubrica','Valor (R$)'],[35,14,50,16]),1):
            c=ws2.cell(row=4,column=col,value=h); ecab(c)
            ws2.column_dimensions[get_column_letter(col)].width=w
        ws2.row_dimensions[4].height=22; linha=5
        for cc, dcc in dfc.groupby('Centro de Custo',sort=False):
            ws2.merge_cells(f'A{linha}:D{linha}')
            c=ws2.cell(row=linha,column=1,value=f'  {cc}')
            c.font=Font(bold=True,color="1F4E79",size=10)
            c.fill=PatternFill("solid",start_color=CC_COR)
            c.alignment=Alignment(horizontal="left",vertical="center")
            ws2.row_dimensions[linha].height=18; linha+=1
            tot={'Provento':0.0,'Desconto':0.0,'Informativo':0.0}
            for _,row in dcc.iterrows():
                cor=cor_tipo(row['Tipo'])
                for col,val in enumerate([cc,row['Tipo'],row['Rubrica'],row['Valor']],1):
                    c=ws2.cell(row=linha,column=col,value=val); edat(c,cor)
                    if col==4: c.number_format='#,##0.00'; c.alignment=Alignment(horizontal="right")
                tot[row['Tipo']]=tot.get(row['Tipo'],0)+row['Valor']; linha+=1
            res=f'Proventos: R$ {tot["Provento"]:,.2f}  |  Descontos: R$ {tot["Desconto"]:,.2f}  |  Informativos: R$ {tot["Informativo"]:,.2f}'
            for col in range(1,5):
                c=ws2.cell(row=linha,column=col,value=res if col==1 else '')
                c.fill=PatternFill("solid",start_color=TOT); c.font=Font(bold=True)
                if col==1: c.alignment=Alignment(horizontal="left",indent=1)
            ws2.row_dimensions[linha].height=16; linha+=2
        ws2.freeze_panes='A5'

    out = io.BytesIO(); wb.save(out); return out.getvalue()

# ─────────────────────────────────────────────
# PARSER: EXTRATO POR FUNCIONARIO
# ─────────────────────────────────────────────

def extrair_rubricas_linha(linha):
    resultado = []
    linha = re.sub(r'(\d{3,6})([A-ZÁÉÍÓÚ])', r'\1 \2', linha)
    linha = re.sub(r'(\d{3,6})(\d+\/)', r'\1 \2', linha)
    for m in re.finditer(r'(\d{1,6})\s+(.{3,50}?)\s+([\d.]+,\d{2})\s*\*?\s*([PD])(?=\s|\d|$)', linha):
        cod = m.group(1)
        nome = re.sub(r'\s+[\d:.,]+$', '', m.group(2)).strip()
        nome = re.sub(r'\s+', ' ', nome).strip()
        try:
            valor = float(m.group(3).replace('.','').replace(',','.'))
        except:
            continue
        if valor > 0 and len(nome) >= 2:
            tipo = 'PROVENTOS' if m.group(4)=='P' else 'DESCONTOS'
            resultado.append((cod, nome, valor, tipo))
    return resultado

def processar_extrato(conteudo_bytes, nome):
    registros = []
    cnpj=razao=comp='N/A'; cc='Sem CC'; func_atual=None; coletando=False
    with pdfplumber.open(io.BytesIO(conteudo_bytes)) as pdf:
        for pagina in pdf.pages:
            for linha in (pagina.extract_text() or "").split('\n'):
                l = linha.strip()
                if not l: continue
                m=re.match(r'Empresa:\s*\d+\s*-\s*(.+)',l,re.I)
                if m: razao=limpar(m.group(1)); continue
                m=re.match(r'CNPJ:\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})',l,re.I)
                if m: cnpj=m.group(1); continue
                m=re.match(r'Compet.ncia:\s*(\d{2}/\d{4})',l,re.I)
                if m: comp=m.group(1); continue
                m=re.match(r'C\.?Custo:\s*(.+)',l,re.I)
                if m: cc=m.group(1).strip(); continue
                m=re.match(r'Empr\.:\s*\d+\s*(.+?)\s+Situa.+?(Trabalhando|Demitido).+?CPF:\s*([\d.\-]+)',l,re.I)
                if m:
                    func_atual={'CNPJ':cnpj,'Razao Social':razao,'Competencia':comp,
                        'Centro de Custo':cc,'Funcionario':m.group(1).strip(),'CPF':m.group(3)}
                    coletando=True; continue
                if re.match(r'^ND:', l, re.I): coletando=False; func_atual=None; continue
                if re.match(r'^(Resumo|L.quido|INSS|Situa|Totais|Base|Cargo:|V.nculo:|Sistema|NF:|Departamento|DEMITIDO|FERIAS DE|Licen|Folha Mensal)', l, re.I): continue
                if func_atual and coletando:
                    for cod,nome_r,valor,tipo in extrair_rubricas_linha(l):
                        registros.append({
                            'Centro de Custo': func_atual['Centro de Custo'],
                            'Funcionario': func_atual['Funcionario'],
                            'CPF': func_atual['CPF'],
                            'Lancamentos': f'    {cod} - {nome_r}',
                            'Tipos': f'  {tipo}',
                            'Valor': valor,
                            'CNPJ': cnpj, 'Razao Social': razao, 'Competencia': comp,
                            '_ordem': 0 if tipo=='PROVENTOS' else 1
                        })
    return registros

def gerar_excel_extrato(registros):
    df = pd.DataFrame(registros)
    df = df.sort_values(['CNPJ','Centro de Custo','Funcionario','_ordem']).drop(columns=['_ordem'])
    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("RESUMO", 0)
    ws.merge_cells('A1:F1'); ws['A1']="RESUMO CONSOLIDADO - EXTRATO MENSAL"
    ws['A1'].font=Font(bold=True,size=14,color="1F4E79"); ws['A1'].alignment=Alignment(horizontal="center")
    for col,(h,w) in enumerate(zip(['CNPJ','Razao Social','Competencia','Total Proventos','Total Descontos','Total Liquido'],[22,42,14,18,18,18]),1):
        c=ws.cell(row=3,column=col,value=h); ecab(c); ws.column_dimensions[get_column_letter(col)].width=w
    linha=4; tp=td=tl=0
    for (cnpj,comp),grp in df.groupby(['CNPJ','Competencia']):
        p=grp[grp.Tipos.str.contains('PROVENTO')].Valor.sum()
        d=grp[grp.Tipos.str.contains('DESCONTO')].Valor.sum()
        l=p-d; tp+=p; td+=d; tl+=l
        for col,val in enumerate([cnpj,grp['Razao Social'].iloc[0],comp,p,d,l],1):
            c=ws.cell(row=linha,column=col,value=val); edat(c)
            if col>=4: c.number_format='R$ #,##0.00'; c.alignment=Alignment(horizontal="right")
        linha+=1
    for col in range(1,7):
        c=ws.cell(row=linha,column=col); c.fill=PatternFill("solid",start_color=TOT); c.font=Font(bold=True)
    ws.merge_cells(f'A{linha}:C{linha}')
    ws.cell(row=linha,column=1,value="TOTAL GERAL").alignment=Alignment(horizontal="center")
    for col,val in zip([4,5,6],[tp,td,tl]):
        c=ws.cell(row=linha,column=col,value=val)
        c.number_format='R$ #,##0.00'; c.font=Font(bold=True); c.alignment=Alignment(horizontal="right")
    ws.freeze_panes='A4'

    for cnpj, dfc in df.groupby('CNPJ'):
        ws2 = wb.create_sheet(re.sub(r'[^\w]','',cnpj)[:31])
        razao=dfc['Razao Social'].iloc[0]; comp=dfc['Competencia'].iloc[0]
        ws2.merge_cells('A1:F1'); ws2['A1']=f'EXTRATO MENSAL - {razao}'
        ws2['A1'].font=Font(bold=True,size=13,color="1F4E79"); ws2['A1'].alignment=Alignment(horizontal="center")
        ws2.merge_cells('A2:F2'); ws2['A2']=f'CNPJ: {cnpj}  |  Competencia: {comp}'
        ws2['A2'].font=Font(italic=True,color="444444"); ws2['A2'].alignment=Alignment(horizontal="center")
        for col,(h,w) in enumerate(zip(['Centro de Custo','Funcionario','CPF','Lancamentos','Tipos','Valor'],[35,30,16,50,14,14]),1):
            c=ws2.cell(row=4,column=col,value=h); ecab(c); ws2.column_dimensions[get_column_letter(col)].width=w
        ws2.row_dimensions[4].height=22; linha=5
        for cc, dcc in dfc.groupby('Centro de Custo',sort=False):
            ws2.merge_cells(f'A{linha}:F{linha}')
            c=ws2.cell(row=linha,column=1,value=f'  {cc}')
            c.font=Font(bold=True,color="1F4E79",size=10)
            c.fill=PatternFill("solid",start_color=CC_COR)
            c.alignment=Alignment(horizontal="left",vertical="center")
            ws2.row_dimensions[linha].height=18; linha+=1
            for _,row in dcc.iterrows():
                cor=PRO if 'PROVENTO' in row['Tipos'] else DES
                for col,val in enumerate([cc,row['Funcionario'],row['CPF'],row['Lancamentos'],row['Tipos'],row['Valor']],1):
                    c=ws2.cell(row=linha,column=col,value=val); edat(c,cor)
                    if col==6: c.number_format='R$ #,##0.00'; c.alignment=Alignment(horizontal="right")
                linha+=1
            tp2=dcc[dcc.Tipos.str.contains('PROVENTO')].Valor.sum()
            td2=dcc[dcc.Tipos.str.contains('DESCONTO')].Valor.sum()
            res=f'Proventos: R$ {tp2:,.2f}  |  Descontos: R$ {td2:,.2f}  |  Liquido: R$ {tp2-td2:,.2f}'
            for col in range(1,7):
                c=ws2.cell(row=linha,column=col,value=res if col==1 else '')
                c.fill=PatternFill("solid",start_color=TOT); c.font=Font(bold=True)
                if col==1: c.alignment=Alignment(horizontal="left",indent=1)
            ws2.row_dimensions[linha].height=16; linha+=2
        ws2.freeze_panes='A5'

    out = io.BytesIO(); wb.save(out); return out.getvalue()

# ─────────────────────────────────────────────
# HANDLER HTTP
# ─────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        if self.path != '/api/converter':
            self.send_response(404); self.end_headers(); return
        try:
            length = int(self.headers.get('Content-Length', 0))
            body = json.loads(self.rfile.read(length))

            resumos = []
            extratos = []

            for arq in body['arquivos']:
                dados = base64.b64decode(arq['conteudo'])
                tipo = detectar_tipo(dados)
                if tipo == 'extrato':
                    extratos.extend(processar_extrato(dados, arq['nome']))
                else:
                    resumos.extend(processar_resumo(dados, arq['nome']))

            resultado = {}
            if resumos:
                resultado['resumo'] = base64.b64encode(gerar_excel_resumo(resumos)).decode()
            if extratos:
                resultado['extrato'] = base64.b64encode(gerar_excel_extrato(extratos)).decode()

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(resultado).encode())

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'erro': str(e)}).encode())
